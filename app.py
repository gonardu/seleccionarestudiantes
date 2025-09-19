import streamlit as st
import json
import os
import hashlib
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ---------- CONFIG ----------
USERS_FILE = "users.json"
DATA_FOLDER = "data"

if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)


# ---------- FUNCIONES ----------
def load_users():
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, "w") as f:
            json.dump({}, f)
    with open(USERS_FILE, "r") as f:
        return json.load(f)


def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=4)


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def login(username, password):
    users = load_users()
    if username in users and users[username]["password"] == hash_password(password):
        return True
    return False


def create_user(username, password, sheet_id="default.xlsx"):
    users = load_users()
    if username in users:
        return False
    users[username] = {
        "password": hash_password(password),
        "sheet_id": sheet_id,
    }
    save_users(users)
    return True


def get_user_file(username):
    users = load_users()
    sheet_id = users[username]["sheet_id"]
    return os.path.join(DATA_FOLDER, sheet_id)


def save_list_to_excel(username, students, results):
    file_path = get_user_file(username)

    # Si no existe el archivo, crear nuevo
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Encabezado inicial (columna A = nombres, fila 1 = fechas)
        ws["A1"] = "Alumno"
        for i, student in enumerate(students, start=2):
            ws.cell(row=i, column=1, value=student)

        wb.save(file_path)

    # Abrir archivo existente
    wb = load_workbook(file_path)
    ws = wb.active

    # Crear nueva columna con fecha actual
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=current_date)

    # Guardar E/X en la columna nueva
    for i, result in enumerate(results, start=2):
        ws.cell(row=i, column=col, value=result)

    wb.save(file_path)


# ---------- INTERFAZ ----------
st.title("📋 Sistema de Asistencia")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = ""

if not st.session_state.logged_in:
    st.subheader("🔑 Iniciar sesión")
    username = st.text_input("Usuario")
    password = st.text_input("Contraseña", type="password")

    if st.button("Ingresar"):
        if login(username, password):
            st.session_state.logged_in = True
            st.session_state.username = username
            st.success("✅ Sesión iniciada")
        else:
            st.error("❌ Usuario o contraseña incorrectos")

    st.subheader("🆕 Crear usuario")
    new_user = st.text_input("Nuevo usuario")
    new_pass = st.text_input("Nueva contraseña", type="password")
    new_file = st.text_input("Nombre de archivo Excel (ej: curso1.xlsx)", "default.xlsx")

    if st.button("Crear usuario"):
        if create_user(new_user, new_pass, new_file):
            st.success("✅ Usuario creado correctamente")
        else:
            st.error("⚠️ El usuario ya existe")

else:
    st.success(f"Sesión iniciada como **{st.session_state.username}**")

    st.subheader("✍️ Ingresar lista de alumnos")
    alumnos_input = st.text_area("Ingrese los nombres separados por coma:", "Juan, Pedro, Maria")
    alumnos = [a.strip() for a in alumnos_input.split(",") if a.strip()]

    if alumnos:
        resultados = []
        for alumno in alumnos:
            opcion = st.radio(f"{alumno}", ["E", "X"], horizontal=True)
            resultados.append(opcion)

        if st.button("💾 Guardar lista en Excel"):
            save_list_to_excel(st.session_state.username, alumnos, resultados)
            st.success("✅ Lista guardada en el archivo Excel")
